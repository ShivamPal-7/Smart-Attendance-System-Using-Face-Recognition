# ============================================================
#  modules/report_generator.py
#  Generate Excel attendance reports using pandas + openpyxl
# ============================================================

import pandas as pd
import os, sys, datetime
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from modules.attendance_manager import get_attendance_report


def generate_excel_report(subject_id=None, from_date=None, to_date=None,
                           class_name=None, save_path=None):
    """
    Pull attendance data and write a formatted Excel file.
    Returns (success: bool, filepath or error message)
    """

    data = get_attendance_report(subject_id, from_date, to_date, class_name)

    if not data:
        return False, "No attendance records found for the selected filters."

    df = pd.DataFrame(data)
    df.rename(columns={
        "name":    "Student Name",
        "roll_no": "Roll No",
        "class":   "Class",
        "subject": "Subject",
        "date":    "Date",
        "time":    "Time",
        "status":  "Status"
    }, inplace=True)

    # Default save path
    if not save_path:
        ts         = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path  = f"Attendance_Report_{ts}.xlsx"

    # ── Write Excel with formatting ───────────────────────────
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance")

        ws = writer.sheets["Attendance"]

        # Header style
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin        = Side(style="thin", color="CCCCCC")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        # Alternate row colors
        light_blue = PatternFill("solid", fgColor="EBF3FB")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = light_blue if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for cell in row:
                cell.fill   = fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        # Color Status column
        status_col = None
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == "Status":
                status_col = col[0].column
                break

        if status_col:
            green = PatternFill("solid", fgColor="D5F5E3")
            red   = PatternFill("solid", fgColor="FADBD8")
            for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
                for cell in row:
                    if cell.value == "Present":
                        cell.fill = green
                        cell.font = Font(color="1E8449", bold=True)
                    else:
                        cell.fill = red
                        cell.font = Font(color="922B21", bold=True)

        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Freeze header
        ws.freeze_panes = "A2"

    print(f"[OK] Report saved: {save_path}")
    return True, save_path


def get_summary_stats(subject_id):
    """Returns summary dict for dashboard display."""
    from modules.attendance_manager import get_attendance_report, get_all_subjects
    from modules.database import execute_query

    data = get_attendance_report(subject_id=subject_id)
    if not data:
        return {"total_classes": 0, "avg_attendance": 0, "below_75": 0}

    df = pd.DataFrame(data)
    total_classes   = df["date"].nunique() if "date" in df else 0
    student_counts  = df[df["status"] == "Present"].groupby("roll_no").size()
    avg             = round(student_counts.mean() / max(total_classes, 1) * 100, 1) if len(student_counts) else 0

    # Students below threshold
    below = execute_query("""
        SELECT COUNT(DISTINCT student_id) AS cnt FROM (
            SELECT student_id,
                   COUNT(CASE WHEN status='Present' THEN 1 END) * 100.0 / COUNT(*) AS pct
            FROM attendance WHERE subject_id=%s
            GROUP BY student_id
            HAVING pct < 75
        ) t
    """, (subject_id,), fetch=True)

    return {
        "total_classes":   total_classes,
        "avg_attendance":  avg,
        "below_75":        below[0]["cnt"] if below else 0
    }
